from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.chart import Reference, BarChart
import datetime


def to_str(s):
    a = 0
    for i in range(1, len(s)):
        if s[i] >= '0' and s[i] <= '9':
            a = a * 10 + ord(s[i]) - ord('0')
    return a


wb = xl.load_workbook("Prices.xlsx")
website = "https://www.flipkart.com/sony-alpha-ilce-7m3-full-frame-mirrorless-camera-body-only-featuring-eye-af-4k-movie-recording/p/itm9e0d891b4eb7f?pid=DLLF7GBGTVVCVHAQ&lid=LSTDLLF7GBGTVVCVHAQVEE3XQ&marketplace=FLIPKART&store=jek%2Fp31%2Ftrv&srno=b_1_1&otracker=hp_omu_Best%2Bof%2BElectronics_1_3.dealCard.OMU_Q5LU1U8PHMK6_3&otracker1=hp_omu_PINNED_neo%2Fmerchandising_Best%2Bof%2BElectronics_NA_dealCard_cc_1_NA_view-all_3&fm=neo%2Fmerchandising&iid=en_NV3oCUjoj9wCPfQ76n65bKsMAmTL8oXMvQz56IZtAMz9uHp1Bb5vC0KBm4tVqLPhpn2%2BcrUx5BUh4T7oX%2BEQ8A%3D%3D&ppt=hp&ppn=homepage&ssid=c0f5mcpa340000001686079275273"
path = "C:\Windows"

options = Options()
options.headless = True
service = Service(executable_path=path)
driver = webdriver.Chrome(service=service, options=options)
driver.get(website)

Title = driver.find_element(by="xpath", value='//h1[@class = "yhB1nd"]').text
Price = driver.find_element(by="xpath", value='//div[@class="_30jeq3 _16Jk6d"]').text
sheet = wb["Sheet1"]
sheet.cell(1, 1).value = f"{Title}"
sheet.cell(1, 1).font = Font("Ariel", bold=True, size=20)
sheet.cell(3, 1).value = "Date: "
sheet.cell(3, 1).font = Font("Ariel", bold=True, size=10)
sheet.cell(3, 2).value = "Price (in Rs): "
sheet.cell(3, 2).font = Font("Ariel", bold=True, size=10)
date = datetime.date.today()
sheet.cell(sheet.max_row + 1, 1).value = f"{date.day}/{date.month}/{date.year}"
sheet.cell(sheet.max_row, 2).value = to_str(Price)  # returns output as a string due to which graph is plotting 0
column = sheet.max_column
values = Reference(sheet, min_row=3, max_row=sheet.max_row, min_col=2, max_col=2)
labels = Reference(sheet, min_row=4, max_row=sheet.max_row, min_col=1, max_col=1)
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(labels)
sheet.add_chart(chart, f"e{column + 1}")
wb.save("Prices.xlsx")
